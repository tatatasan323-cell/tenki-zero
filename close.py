# -*- coding: utf-8 -*-
"""tenki-zero ─ 月次締め（転記ゼロ）

使い方:
    python close.py 2026-05          … コマンドで締める
    python app.py                    … ブラウザの受付画面から使う（推奨）

inbox/ に放り込まれた〈売上・経費・勤怠〉を読み、SQLite（数字の泉）を作り直し、
out/YYYY-MM/ に月次帳票一式（CSV/HTML/Excel）を出力する。

設計原則:
  - 源流で一度だけ記録（数字の住所は1か所）→ 転記は構造的に存在しない
  - 判断は人（masters/ の費目・ルール）、適用は機械
  - 出力は専門システム（会計ソフト・税理士）に渡す直前まで
  - 現実データは汚い前提: SJIS/UTF-8自動判定・列名ゆれ吸収・¥カンマ除去・重複自動排除
  - CSVだけなら標準ライブラリのみで動く。Excel/PDFの取込とExcel帳票は“借り物”
    （openpyxl / PyMuPDF ― requirements.txt）を入れた時だけ有効になる
"""
import os, io, csv, re, sys, sqlite3, datetime

# 借り物の部品はlib/に同梱済み(=増築済みでお届け)。venv等に自前導入があればそちらでも動く
sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), "lib"))
try:
    import openpyxl                     # 借り物① Excel読み書き
except ImportError:
    openpyxl = None
try:
    import fitz                         # 借り物② PDF読み取り(PyMuPDF)
except ImportError:                     # 32bit環境や他OSでは同梱版が動かない場合あり→CSV機能のみで続行
    fitz = None

BASE = os.path.dirname(os.path.abspath(__file__))
J = os.path.join

# ---------- 汚いデータ吸収 ----------
def smart_read(path):
    raw = open(path, "rb").read()
    if raw[:3] == b"\xef\xbb\xbf":
        return raw[3:].decode("utf-8")
    for enc in ("utf-8", "cp932", "euc-jp"):
        try:
            return raw.decode(enc)
        except UnicodeDecodeError:
            continue
    return raw.decode("utf-8", errors="replace")

def rows_of(path, warnings):
    """CSV/TSV/Excel/PDF を同じ「行の表」に揃える"""
    ext = os.path.splitext(path)[1].lower()
    if ext in (".xlsx", ".xlsm"):
        if not openpyxl:
            warnings.append("openpyxl未導入のためExcelをスキップ: " + os.path.basename(path))
            return []
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.worksheets[0]
        return [[("" if c is None else str(c)) for c in row]
                for row in ws.iter_rows(values_only=True)
                if any(c is not None and str(c).strip() for c in row)]
    if ext == ".pdf":
        if not fitz:
            warnings.append("PyMuPDF未導入のためPDFをスキップ: " + os.path.basename(path))
            return []
        doc = fitz.open(path)
        lines = []
        for page in doc:
            lines += [l for l in page.get_text().splitlines() if l.strip()]
        rows = [re.split(r"[,\t]|\s{2,}", l.strip()) for l in lines]
        if not any(len(r) >= 2 for r in rows):
            warnings.append("PDFから表を読み取れませんでした（AIに『CSVにして』と頼むのが早道です）: "
                            + os.path.basename(path))
        return rows
    text = smart_read(path).replace("\r\n", "\n").replace("\r", "\n")
    delim = "\t" if text.count("\t") > text.count(",") else ","
    return [r for r in csv.reader(io.StringIO(text), delimiter=delim) if any(c.strip() for c in r)]

def num(s):
    t = re.sub(r"[^0-9.\-]", "", str(s))
    try:
        return float(t)
    except ValueError:
        return 0.0

def dt(s):
    m = re.match(r"(\d{4})\D(\d{1,2})\D(\d{1,2})", str(s).strip())
    return "%04d-%02d-%02d" % (int(m.group(1)), int(m.group(2)), int(m.group(3))) if m else None

ALIAS = {
    "date": ["日付", "請求日", "取引日", "営業日", "伝票日付", "date"],
    "store": ["店舗", "店名", "所属", "store"],
    "amount": ["売上", "金額", "請求額", "税込金額", "金額(税込)", "amount", "total"],
    "vendor": ["取引先", "仕入先", "請求元", "支払先", "vendor"],
    "item": ["品目", "内容", "摘要", "品名", "item"],
    "emp": ["従業員コード", "社員コード", "emp"],
    "minutes": ["実労働時間(分)", "労働時間(分)", "実労働分", "minutes"],
}
def header_map(hdr):
    m = {}
    for i, h in enumerate(hdr):
        k = str(h).strip()
        for field, names in ALIAS.items():
            if k in names:
                m[field] = i
    return m

def read_table(path, need, warnings):
    rows = rows_of(path, warnings)
    for hi in range(min(10, len(rows))):
        m = header_map(rows[hi])
        if all(k in m for k in need):
            out = []
            for r in rows[hi + 1:]:
                rec = {k: (str(r[i]).strip() if i < len(r) else "") for k, i in m.items()}
                out.append(rec)
            return out
    if rows:
        warnings.append("ヘッダを見つけられずスキップ: " + os.path.basename(path))
    return []

def load_master(name):
    p = J(BASE, "masters", name)
    return [r for r in csv.reader(io.StringIO(smart_read(p))) if any(c.strip() for c in r)][1:]

def vendor_from_filename(path):
    return re.sub(r"(_請求書|請求書?|_invoice|invoice)$", "",
                  os.path.splitext(os.path.basename(path))[0], flags=re.I)

# ---------- 取込 → SQLite（重複は自動排除・毎回作り直し＝冪等） ----------
def build_db(month, warnings):
    items = {c: (n, g) for c, n, g in load_master("items.csv")}
    rules = load_master("rules.csv")
    emps = {r[0]: r for r in load_master("employees.csv")}

    con = sqlite3.connect(J(BASE, "out", "tenki.sqlite"))
    con.executescript("""
      DROP TABLE IF EXISTS sales;      CREATE TABLE sales(date TEXT, store TEXT, amount REAL, src TEXT);
      DROP TABLE IF EXISTS expenses;   CREATE TABLE expenses(date TEXT, vendor TEXT, item TEXT, amount REAL,
                                          item_code TEXT, item_name TEXT, src TEXT);
      DROP TABLE IF EXISTS attendance; CREATE TABLE attendance(date TEXT, emp TEXT, name TEXT, store TEXT,
                                          minutes REAL, src TEXT);
    """)
    dedup = {"sales_replaced": 0, "expenses_skipped": 0, "attendance_replaced": 0}
    month_of = lambda d: d and d[:7] == month
    def scan(sub):
        d = J(BASE, "inbox", sub)
        fs = [f for f in os.listdir(d)
              if not f.startswith(".") and "お読みください" not in f
              and f.lower().endswith((".csv", ".tsv", ".xlsx", ".xlsm", ".pdf"))]
        # 「後にアップロードした方が勝つ」を字義通りにする: 受付に置かれた時刻順に読む(同時刻は名前順)
        return [J(d, f) for f in sorted(fs, key=lambda f: (os.path.getmtime(J(d, f)), f))]

    # 売上: 同一(日付,店舗)は後から読んだものが優先(再アップロード=上書き)
    sales = {}
    for p in scan("sales"):
        for r in read_table(p, ("date", "amount"), warnings):
            d = dt(r.get("date"))
            if not (month_of(d) and num(r.get("amount"))):
                continue
            store = r.get("store") or vendor_from_filename(p)
            key = (d, store)
            if key in sales:
                dedup["sales_replaced"] += 1
            sales[key] = (num(r["amount"]), os.path.basename(p))
    for (d, s), (a, src) in sales.items():
        con.execute("INSERT INTO sales VALUES(?,?,?,?)", (d, s, a, src))

    # 経費: 完全一致行(日付,取引先,品目,金額)の重複はスキップ。合計/小計行も除外
    seen = set()
    for p in scan("expenses"):
        fallback = vendor_from_filename(p)
        for r in read_table(p, ("date", "amount"), warnings):
            d = dt(r.get("date"))
            item = r.get("item", "")
            if re.search(r"合計|小計|total", item, re.I):
                continue
            if not (month_of(d) and num(r.get("amount"))):
                continue
            vendor = r.get("vendor") or fallback
            key = (d, vendor, item, round(num(r["amount"])))
            if key in seen:
                dedup["expenses_skipped"] += 1
                continue
            seen.add(key)
            code = "MISC"
            haystack = (vendor + " " + item).lower()   # 取引先＋品目の両方で仕分け
            for pat, c in rules:
                if pat.lower() in haystack:
                    code = c
                    break
            con.execute("INSERT INTO expenses VALUES(?,?,?,?,?,?,?)",
                        (d, vendor, item, num(r["amount"]), code,
                         items.get(code, (code, ""))[0], os.path.basename(p)))

    # 勤怠: 同一(日付,従業員)は後勝ち
    att = {}
    for p in scan("attendance"):
        for r in read_table(p, ("date", "emp", "minutes"), warnings):
            d = dt(r.get("date"))
            if not month_of(d):
                continue
            key = (d, r["emp"])
            if key in att:
                dedup["attendance_replaced"] += 1
            e = emps.get(r["emp"])
            att[key] = (r.get("name", e[1] if e else ""), e[2] if e else r.get("store", ""),
                        num(r["minutes"]), os.path.basename(p))
    for (d, emp), (name, store, mins, src) in att.items():
        con.execute("INSERT INTO attendance VALUES(?,?,?,?,?,?)", (d, emp, name, store, mins, src))

    con.commit()
    return con, items, emps, dedup

# ---------- 帳票 ----------
def yen(n): return "¥{:,}".format(round(n))
def wcsv(path, header, rows):
    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        w.writerow(header)
        w.writerows(rows)

def payroll(con, emps):
    out = []
    for code, (c, name, dept, wtype, rate) in sorted(emps.items()):
        mins = con.execute("SELECT COALESCE(SUM(minutes),0), COUNT(DISTINCT date) FROM attendance WHERE emp=?",
                           (code,)).fetchone()
        base = int(rate) if wtype == "月給" else round(mins[0] / 60 * int(rate))
        out.append([code, name, dept, wtype, rate, mins[1], round(mins[0] / 60, 1), base])
    return out

def build_matrix(con, items, pay, by_store):
    """部門別損益マトリクス: 行=勘定科目 / 列=各店舗＋本部＋全社合計。
    仕入(COGS)は売上比で店舗按分、その他は品目/取引先/所属から部門を判定。"""
    stores = [s for s, _ in by_store]                # 売上のある店(降順)
    sales_by = dict(by_store)
    stotal = sum(sales_by.values()) or 1
    cols = stores + ["本部", "全社合計"]
    def blank():
        return dict((c, 0) for c in cols)
    def dept_of(item, vendor, code):
        t = (item or "") + " " + (vendor or "")
        for s in stores:
            if s in t:
                return s
        return "本部"                                # 店舗名が無ければ本部(管理部門)扱い
    rev = blank()
    for s, a in by_store:
        rev[s] = a
    rev["全社合計"] = sum(sales_by.values())
    cogs_total = 0
    sga = []                                         # [(item_name, {col:amt})] 出現順
    sga_map = {}
    for d, v, it, a, code, iname in con.execute(
            "SELECT date,vendor,item,amount,item_code,item_name FROM expenses"):
        if code == "COGS":
            cogs_total += a
            continue
        row = sga_map.get(iname)
        if row is None:
            row = blank(); sga_map[iname] = row; sga.append((iname, row))
        dep = dept_of(it, v, code)
        row[dep] += a; row["全社合計"] += a
    pr = sga_map.get("給与手当")
    if pr is None:
        pr = blank(); sga_map["給与手当"] = pr; sga.insert(0, ("給与手当", pr))
    for r in pay:
        pr[r[2]] = pr.get(r[2], 0) + r[7]; pr["全社合計"] += r[7]
    cogs = blank()
    for s in stores:
        cogs[s] = round(cogs_total * sales_by[s] / stotal)
    cogs["全社合計"] = cogs_total
    gross = dict((c, rev[c] - cogs[c]) for c in cols)
    sga = sorted(sga, key=lambda x: (x[0] != "給与手当", -x[1]["全社合計"]))
    sga_tot = dict((c, sum(r[c] for _, r in sga)) for c in cols)
    op = dict((c, gross[c] - sga_tot[c]) for c in cols)
    R = lambda d: [round(d[c]) for c in cols]
    rows = [["売上高"] + R(rev), ["売上原価(仕入・按分)"] + R(cogs), ["売上総利益"] + R(gross)]
    rows += [[n] + R(r) for n, r in sga]
    rows += [["販管費計"] + R(sga_tot), ["営業利益"] + R(op)]
    return cols, rows

def close(month):
    warnings = []
    outdir = J(BASE, "out", month)
    os.makedirs(outdir, exist_ok=True)
    con, items, emps, dedup = build_db(month, warnings)

    sales_total = con.execute("SELECT COALESCE(SUM(amount),0) FROM sales").fetchone()[0]
    if not sales_total:
        return {"ok": False, "error": "%s の売上データが見つかりません。inbox/sales/ を確認してください。" % month,
                "warnings": warnings}
    by_store = con.execute("SELECT store, SUM(amount) FROM sales GROUP BY store ORDER BY 2 DESC").fetchall()
    by_day = con.execute("SELECT date, SUM(amount) FROM sales GROUP BY date ORDER BY 1").fetchall()
    exp = con.execute("SELECT item_code, item_name, SUM(amount), COUNT(*) FROM expenses GROUP BY 1,2 ORDER BY 3 DESC").fetchall()
    misc = con.execute("SELECT date, vendor, item, amount FROM expenses WHERE item_code='MISC'").fetchall()
    pay = payroll(con, emps)
    pay_total = sum(r[7] for r in pay)
    cogs = sum(a for c, n, a, k in exp if items.get(c, ("", ""))[1] == "売上原価")
    sga_inv = [(n, a) for c, n, a, k in exp if items.get(c, ("", ""))[1] != "売上原価"]
    sga_total = sum(a for _, a in sga_inv) + pay_total
    gross = sales_total - cogs
    op = gross - sga_total

    pl_rows = [["売上高", round(sales_total), ""],
               ["売上原価(仕入高)", round(cogs), "経費データより"],
               ["売上総利益", round(gross), "粗利率 {:.1f}%".format(gross / sales_total * 100)],
               ["給与手当", pay_total, "勤怠データより"]]
    pl_rows += [[n, round(a), "請求書より"] for n, a in sorted(sga_inv, key=lambda x: -x[1])]
    pl_rows += [["販管費計", round(sga_total), ""],
                ["営業利益", round(op), "営業利益率 {:.1f}%".format(op / sales_total * 100)]]
    wcsv(J(outdir, "01_月次損益.csv"), ["科目", "金額", "備考"], pl_rows)

    exp_total = cogs + sum(a for _, a in sga_inv)
    exp_rows = [[n, round(a), k, round(a / exp_total * 100, 1)] for c, n, a, k in exp]
    wcsv(J(outdir, "02_経費内訳_費目別.csv"), ["費目", "金額", "明細数", "構成比%"], exp_rows)
    store_rows = [[s, round(a), round(a / sales_total * 100, 1)] for s, a in by_store]
    wcsv(J(outdir, "03_売上計上_店舗別.csv"), ["店舗", "売上", "構成比%"], store_rows)
    att = con.execute("SELECT emp, name, store, COUNT(DISTINCT date), SUM(minutes) FROM attendance GROUP BY 1,2,3 ORDER BY 1").fetchall()
    wcsv(J(outdir, "04_勤怠集計.csv"), ["従業員コード", "氏名", "所属", "出勤日数", "実労働時間(h)"],
         [[e, n, s, d, round(m / 60, 1)] for e, n, s, d, m in att])
    wcsv(J(outdir, "05_給与もと資料.csv"),
         ["従業員コード", "氏名", "所属", "給与区分", "単価", "出勤日数", "実労働時間(h)", "基本支給額(もと)"], pay)
    dept_pay = {}
    for r in pay:
        dept_pay[r[2]] = dept_pay.get(r[2], 0) + r[7]
    dept_rows = [[s, round(a), dept_pay.get(s, 0), round(dept_pay.get(s, 0) / a * 100, 1)] for s, a in by_store]
    wcsv(J(outdir, "06_部門別実績.csv"), ["部門", "売上", "人件費(もと)", "人件費率%"], dept_rows)

    j = [[d, "売上高", "", round(a), "売上計上 " + s, s] for d, s, a in
         con.execute("SELECT date, store, amount FROM sales ORDER BY date").fetchall()]
    j += [[d, i2, c, round(a), (v + " " + it).strip(), ""] for d, v, it, a, c, i2 in
          con.execute("SELECT date, vendor, item, amount, item_code, item_name FROM expenses ORDER BY date").fetchall()]
    wcsv(J(outdir, "08_税理士パック_仕訳候補.csv"), ["日付", "科目", "費目コード", "金額", "摘要", "部門"], j)
    wcsv(J(outdir, "09_月次収支サマリ.csv"), ["区分", "金額"],
         [["収入(売上)", round(sales_total)], ["支出(経費+給与もと)", round(cogs + sga_total)],
          ["差引", round(sales_total - cogs - sga_total)]])

    mcols, mrows = build_matrix(con, items, pay, by_store)
    wcsv(J(outdir, "11_部門別損益マトリクス.csv"), ["勘定科目"] + mcols, mrows)

    try:
        holidays = set(r[0] for r in load_master("holidays.csv"))
    except Exception:
        holidays = set()
    dashboard(outdir, month, sales_total, gross, op, pay_total, cogs, by_store, by_day, exp, items, misc, dept_pay,
              (mcols, mrows), pay_total, holidays)
    if openpyxl:
        excel_report(outdir, month, pl_rows, exp_rows, store_rows, dept_rows, (mcols, mrows))
    else:
        warnings.append("openpyxl未導入のためExcel帳票(10_月次報告.xlsx)は省略（requirements.txt参照）")

    outputs = sorted(f for f in os.listdir(outdir) if not f.startswith("."))
    return {"ok": True, "month": month, "sales": round(sales_total), "gross": round(gross),
            "gross_rate": round(gross / sales_total * 100, 1), "op": round(op),
            "op_rate": round(op / sales_total * 100, 1), "misc": len(misc),
            "dedup": dedup, "warnings": warnings, "outdir": outdir, "outputs": outputs}

def excel_report(outdir, month, pl_rows, exp_rows, store_rows, dept_rows, matrix):
    """役員会議・銀行提出にそのまま使えるExcelブック（借り物openpyxl使用）"""
    wb = openpyxl.Workbook()
    bold = openpyxl.styles.Font(bold=True)
    fill = openpyxl.styles.PatternFill("solid", fgColor="1F4E79")
    white = openpyxl.styles.Font(bold=True, color="FFFFFF")
    def sheet(name, header, rows, money_cols):
        ws = wb.create_sheet(name)
        ws.append(header)
        for c in ws[1]:
            c.font = white
            c.fill = fill
        for r in rows:
            ws.append(r)
        for col in money_cols:
            for cell in ws[openpyxl.utils.get_column_letter(col)][1:]:
                cell.number_format = "#,##0"
        for i, _ in enumerate(header, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 16
        return ws
    wb.remove(wb.active)
    mcols, mrows = matrix
    ws = sheet("部門別損益マトリクス", ["勘定科目"] + mcols, mrows, list(range(2, len(mcols) + 2)))
    for row in ws.iter_rows(min_row=2):
        if row[0].value in ("売上総利益", "営業利益"):
            for c in row:
                c.font = bold
    sheet("月次損益", ["科目", "金額", "備考"], pl_rows, [2])
    sheet("経費内訳", ["費目", "金額", "明細数", "構成比%"], exp_rows, [2])
    sheet("店舗別売上", ["店舗", "売上", "構成比%"], store_rows, [2])
    sheet("部門別実績", ["部門", "売上", "人件費(もと)", "人件費率%"], dept_rows, [2, 3])
    gray = openpyxl.styles.Font(color="8296AD", size=9)
    for ws in wb.worksheets:
        ws.append([])
        ws.append(["tenki-zero 自動出力 ／ 制作：AI内製化工房 MITA ── 会計・税務の最終判断は税理士等の専門家にご相談ください"])
        ws.cell(ws.max_row, 1).font = gray
    wb.save(J(outdir, "10_月次報告_%s.xlsx" % month))

# ---------- 役員ダッシュボード ----------
def svg_bars(pairs, color, fmt=lambda v: "¥" + format(round(v / 1000), ",") + "k", label_every=1, show_val=True):
    W, H, pl, pb, pt = 640, 220, 10, 30, 22
    if not pairs:
        return ""
    mx = max(v for _, v in pairs)
    n = len(pairs)
    gap = (W - pl * 2) / n
    bw = min(56, gap * 0.62)
    b = ""
    for i, (lab, v) in enumerate(pairs):
        x = pl + gap * i + (gap - bw) / 2
        h = (H - pt - pb) * (v / mx)
        y = pt + (H - pt - pb) - h
        b += '<rect x="%.1f" y="%.1f" width="%.1f" height="%.1f" rx="3" fill="%s"/>' % (x, y, bw, h, color)
        if show_val:
            b += '<text x="%.1f" y="%.1f" text-anchor="middle" font-size="10.5" fill="#c3d2e4">%s</text>' % (x + bw / 2, y - 4, fmt(v))
        if i % label_every == 0:
            b += '<text x="%.1f" y="%d" text-anchor="middle" font-size="11" fill="#aebccd">%s</text>' % (x + bw / 2, H - pb + 15, lab)
    return '<svg viewBox="0 0 %d %d" width="100%%" style="max-width:680px">%s</svg>' % (W, H, b)

def matrix_table(mcols, mrows):
    """部門別損益マトリクスをHTML表に。売上総利益・営業利益は強調、マイナスは赤。"""
    def cell(v, emph):
        neg = " style=\"color:#ff8a8a\"" if v < 0 else ""
        return "<td%s>%s</td>" % (neg, ("<b>%s</b>" % yen(v)) if emph else yen(v))
    th = "".join("<th>%s</th>" % c for c in mcols)
    body = ""
    for r in mrows:
        emph = r[0] in ("売上総利益", "営業利益", "売上高")
        cls = ' class="hl"' if r[0] in ("売上総利益", "営業利益") else ""
        body += "<tr%s><td class=\"l\">%s</td>%s</tr>" % (cls, r[0], "".join(cell(v, emph) for v in r[1:]))
    return ('<div style="overflow-x:auto"><table class="mx"><tr><th class="l">勘定科目</th>%s</tr>%s</table></div>'
            '<p class="muted">※仕入(売上原価)は売上比で店舗按分。本部＝管理部門(売上なし＝費用のみ)。金額は自動集計・転記なし。</p>'
            % (th, body))

WDN = ["月", "火", "水", "木", "金", "土", "日"]
PALETTE = ["#5aa2e6", "#37c39a", "#f5a524", "#ff6b81", "#9b7ede", "#4fc3e0", "#e0a24a",
           "#7fca6b", "#e07aa0", "#c9863f", "#6f9fce", "#d9b34a", "#8fd0ff", "#ff9f7a"]

def svg_daily(rows, holidays):
    """日次売上。土=水色/日=赤/祝=橙/平日=青で色分け。ホバーで日付・曜日・金額。"""
    W, H, pl, pb, pt, pr = 720, 220, 48, 30, 16, 8
    plotH = H - pt - pb
    mx = max((a for _, a in rows), default=1) or 1
    n = len(rows) or 1
    gap = (W - pl - pr) / n
    bw = min(18, gap * 0.72)
    COL = {"w": "#5aa2e6", "sat": "#3fb7d6", "sun": "#ff6b81", "hol": "#f5a524"}
    LCOL = {"w": "#9fb0c7", "sat": "#7fd6ec", "sun": "#ff9db0", "hol": "#ffca7a"}
    g = ""
    for i in range(3):
        y = pt + plotH * (1 - i / 2)
        g += ('<line x1="%d" y1="%.1f" x2="%d" y2="%.1f" stroke="rgba(255,255,255,.08)"/>'
              '<text x="%d" y="%.1f" text-anchor="end" font-size="10" fill="#8296ad">%s</text>'
              % (pl, y, W - pr, y, pl - 6, y + 4, "¥" + format(round(mx * i / 2 / 1000), ",") + "k"))
    for i, (d, a) in enumerate(rows):
        wd = datetime.date(*map(int, d.split("-"))).weekday()
        t = "hol" if d in holidays else ("sat" if wd == 5 else ("sun" if wd == 6 else "w"))
        x = pl + gap * i + (gap - bw) / 2
        h = plotH * (a / mx)
        y = pt + plotH - h
        g += ('<rect x="%.1f" y="%.1f" width="%.1f" height="%.1f" rx="2" fill="%s"><title>%s(%s) %s</title></rect>'
              '<text x="%.1f" y="%d" text-anchor="middle" font-size="8.5" fill="%s">%d</text>'
              % (x, y, bw, h, COL[t], d, WDN[wd], yen(a), x + bw / 2, H - pb + 12, LCOL[t], int(d[8:])))
    return '<svg viewBox="0 0 %d %d" width="100%%" style="max-width:100%%">%s</svg>' % (W, H, g)

def svg_hstack(segs):
    """横方向の積み上げ棒（1本を費目で色分け）。"""
    W, H = 720, 40
    total = sum(a for _, a, _ in segs) or 1
    x = 0.0
    g = ""
    for name, a, color in segs:
        w = W * a / total
        g += ('<rect x="%.1f" y="6" width="%.1f" height="28" fill="%s"><title>%s %s（%.1f%%）</title></rect>'
              % (x, max(0.5, w), color, name, yen(a), a / total * 100))
        x += w
    return '<svg viewBox="0 0 %d %d" width="100%%" style="max-width:100%%">%s</svg>' % (W, H, g)

def legend(pairs):
    return '<div class="lgd">%s</div>' % "".join(
        '<span><i style="background:%s"></i>%s</span>' % (c, t) for t, c in pairs)

def dashboard(outdir, month, sales, gross, op, pay_total, cogs, by_store, by_day, exp, items, misc, dept_pay,
              matrix, pay_tot, holidays):
    wsum = {}
    for d, a in by_day:
        wsum[WDN[datetime.date(*map(int, d.split("-"))).weekday()]] = \
            wsum.get(WDN[datetime.date(*map(int, d.split("-"))).weekday()], 0) + a
    top_wd = max(wsum, key=wsum.get)
    top_store, top_amt = by_store[0]
    low_store, low_amt = by_store[-1]
    sga_exp = [(n, a) for c, n, a, k in exp if items.get(c, ("", ""))[1] != "売上原価"]
    comments = [
        "売上トップは %s（%s・構成比 %.1f%%）。最下位 %s（%s）との差は %.1f倍。" %
        (top_store, yen(top_amt), top_amt / sales * 100, low_store, yen(low_amt), top_amt / max(low_amt, 1)),
        "曜日では %s曜が最大（%s）。仕込み・シフトは%s曜に厚く。" % (top_wd, yen(wsum[top_wd]), top_wd),
        "原価率 %.1f%%（仕入 %s）／粗利率 %.1f%%／人件費率 %.1f%%。" %
        (cogs / sales * 100, yen(cogs), gross / sales * 100, pay_total / sales * 100),
        "販管費で大きいのは %s。" % "、".join("%s %s" % (n, yen(a)) for n, a in sorted(sga_exp, key=lambda x: -x[1])[:3]),
    ]
    comments.append("未分類の経費が %d件 ―― masters/rules.csv にルール追加を。" % len(misc) if misc
                    else "経費の未分類は 0件 ―― 仕分けルールは健在です。")
    kpi = (("売上高", yen(sales)), ("売上総利益", yen(gross)), ("営業利益", yen(op)),
           ("原価率", "%.1f%%" % (cogs / sales * 100)), ("人件費率", "%.1f%%" % (pay_total / sales * 100)))
    kpis = "".join('<div class="kpi"><div class="v">%s</div><div class="k">%s</div></div>' % (v, k) for k, v in kpi)

    daily = svg_daily(by_day, holidays)
    day_leg = legend([("平日", "#5aa2e6"), ("土", "#3fb7d6"), ("日", "#ff6b81"), ("祝", "#f5a524")])
    cost = [("仕入高", cogs), ("給与手当", pay_tot)] + sga_exp
    cost = [(n, a) for n, a in cost if a > 0]
    cost.sort(key=lambda x: -x[1])
    tot_cost = sum(a for _, a in cost) or 1
    segs = [(n, a, PALETTE[i % len(PALETTE)]) for i, (n, a) in enumerate(cost)]
    hstack = svg_hstack(segs)
    exp_leg = '<div class="lgd">%s</div>' % "".join(
        '<span><i style="background:%s"></i>%s ¥%s（%.1f%%）</span>' % (c, n, format(round(a), ","), a / tot_cost * 100)
        for n, a, c in segs)
    stores = svg_bars(by_store, "#37c39a")
    mtable = matrix_table(*matrix)
    lis = "".join("<li>%s</li>" % c for c in comments)

    html = """<!DOCTYPE html><html lang="ja"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>役員ダッシュボード %(m)s</title><style>
body{margin:0;font-family:"Segoe UI","Yu Gothic UI",Meiryo,sans-serif;color:#e9f0fa;line-height:1.6;
 background:linear-gradient(165deg,#0a1020,#0c162b 55%%,#0a1224);min-height:100vh}
.wrap{max-width:1360px;margin:0 auto;padding:20px 18px}
h1{font-size:1.35rem;margin:.2em 0 .6em;background:linear-gradient(112deg,#fff,#8fd0ff 60%%,#ffd67a);-webkit-background-clip:text;color:transparent}
.card{background:rgba(16,25,44,.66);border:1px solid rgba(255,255,255,.1);border-radius:16px;padding:14px 18px;margin:0 0 14px}
h3{font-size:.98rem;color:#eaf3ff;border-left:4px solid #5aa2e6;padding-left:10px;margin:.1em 0 .5em}
.kpis{display:flex;gap:12px;flex-wrap:wrap}
.kpi{flex:1;min-width:120px;background:linear-gradient(158deg,rgba(74,150,235,.22),rgba(13,22,42,.66));
 border:1px solid rgba(255,255,255,.1);border-radius:13px;padding:11px 14px}
.kpi .v{font-size:1.25rem;font-weight:800;color:#eaf4ff}.kpi .k{font-size:.72rem;color:#9fb2c9}
li{margin:.3em 0;font-size:.9rem}.muted{color:#8ea1b8;font-size:.8rem}
.grid2{display:grid;grid-template-columns:minmax(0,1.5fr) minmax(0,1fr);gap:16px;align-items:start}
@media(max-width:920px){.grid2{grid-template-columns:1fr}}
.lgd{display:flex;flex-wrap:wrap;gap:6px 14px;font-size:.76rem;color:#aebccd;margin:6px 0}
.lgd i{display:inline-block;width:11px;height:11px;border-radius:3px;margin-right:4px;vertical-align:-1px}
table.mx{border-collapse:collapse;width:100%%;font-size:.82rem;min-width:600px}
table.mx th,table.mx td{border:1px solid rgba(255,255,255,.1);padding:5px 8px;text-align:right;color:#cdd9e8;white-space:nowrap}
table.mx th{background:rgba(90,160,240,.16);color:#eaf3ff}
table.mx td.l,table.mx th.l{text-align:left;position:sticky;left:0;background:#0e1a30}
table.mx tr.hl td{background:rgba(55,195,154,.12);color:#eaffef}
table.mx tr:last-child td{border-top:2px solid rgba(255,255,255,.25)}
</style></head><body><div class="wrap">
<h1>役員ダッシュボード ─ %(m)s</h1>
<div class="card"><div class="kpis">%(kpis)s</div></div>
<div class="grid2">
  <div>
    <div class="card"><h3>部門別 損益マトリクス（全社／店舗／管理部門）</h3>%(mtable)s</div>
    <div class="card"><h3>分析コメント（自動生成）</h3><ul>%(lis)s</ul></div>
  </div>
  <div>
    <div class="card"><h3>日次売上（土日祝を色分け）</h3>%(day_leg)s%(daily)s</div>
    <div class="card"><h3>経費・費用の内訳（積み上げ）</h3>%(hstack)s%(exp_leg)s</div>
    <div class="card"><h3>店舗別 売上</h3>%(stores)s</div>
  </div>
</div>
<p class="muted" style="border-top:1px solid rgba(255,255,255,.08);padding-top:12px;margin-top:18px">
<b>tenki-zero ─ 役員ダッシュボード</b> ／ 制作：AI内製化工房 MITA　｜　公式チャンネル：https://note.com/ai_naiseika<br>
自動出力 ─ 数字の出所: inbox/（売上・経費・勤怠）。転記は行われていません。<br>
ご利用は自己責任で。会計・税務の最終判断は税理士等の専門家にご相談ください。</p>
</div></body></html>""" % dict(m=month, kpis=kpis, lis=lis, stores=stores, daily=daily, day_leg=day_leg,
                                 hstack=hstack, exp_leg=exp_leg, mtable=mtable)
    with open(J(outdir, "07_役員ダッシュボード.html"), "w", encoding="utf-8") as f:
        f.write(html)

if __name__ == "__main__":
    if len(sys.argv) != 2 or not re.match(r"^\d{4}-\d{2}$", sys.argv[1]):
        print("使い方: python close.py YYYY-MM   例: python close.py 2026-05")
        sys.exit(1)
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    r = close(sys.argv[1])
    if not r["ok"]:
        print("エラー:", r["error"])
        sys.exit(1)
    print("=== %s 月次締め 完了 ===" % r["month"])
    print("売上 %s ／ 粗利 %s(%.1f%%) ／ 営業利益 %s(%.1f%%)"
          % (yen(r["sales"]), yen(r["gross"]), r["gross_rate"], yen(r["op"]), r["op_rate"]))
    d = r["dedup"]
    if any(d.values()):
        print("重複排除: 売上上書き%d件 / 経費スキップ%d件 / 勤怠上書き%d件"
              % (d["sales_replaced"], d["expenses_skipped"], d["attendance_replaced"]))
    print("未分類: %d件%s" % (r["misc"], " → masters/rules.csv にルール追加を" if r["misc"] else "（なし）"))
    for w in r["warnings"]:
        print("注意:", w)
    print("出力先:", r["outdir"])
